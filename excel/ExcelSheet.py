from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# TODO: do more design for this class.

class ExcelSheet:
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename)
        self.ws = self.wb.active

    def get_cell(self, row, col):
        col_letter = get_column_letter(col)
        cell = self.ws['{}{}'.format(col_letter, row)]
        return cell.value

    def set_cell(self, row, col, value):
        col_letter = get_column_letter(col)
        self.ws['{}{}'.format(col_letter, row)] = value
        self.wb.save(self.filename)
        
        
if __name__ == '__main__':
    # Create a new Excel file with some data
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name'
    ws['B1'] = 'Age'
    ws['A2'] = 'Alice'
    ws['B2'] = 25
    ws['A3'] = 'Bob'
    ws['B3'] = 30
    wb.save('example.xlsx')

    # Create an instance of the ExcelSheet class for the file we just created
    sheet = ExcelSheet('example.xlsx')

    # Get the value of a cell
    name = sheet.get_cell(2, 1)
    print(name)  # Output: Alice

    # Set the value of a cell
    sheet.set_cell(3, 2, 35)

    # Get the new value of the cell
    age = sheet.get_cell(3, 2)
    print(age)  # Output: 35